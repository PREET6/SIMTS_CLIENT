import os
from datetime import datetime
from pathlib import Path

from flask import Blueprint, current_app, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from extensions import db
from models import Admin, AuditLog, Certificate, ContactMessage, Course, Marksheet, Student
from security import admin_required, audit
from services.backup import create_backup

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')


def protect(view):
    return login_required(admin_required(view))


def clean(value, max_length):
    return (value or '').strip()[:max_length]


def pdf_signature(file_storage):
    try:
        pos = file_storage.stream.tell()
        file_storage.stream.seek(0)
        header = file_storage.stream.read(5)
        file_storage.stream.seek(pos)
        return header == b'%PDF-'
    except Exception:
        return False


# ============================================================
# DASHBOARD
# ============================================================

@admin_bp.route('/')
@protect
def dashboard():
    return render_template(
        'admin/dashboard.html',
        student_count=Student.query.count(),
        course_count=Course.query.count(),
        certificate_count=Certificate.query.count(),
        marksheet_count=Marksheet.query.count(),
        message_count=ContactMessage.query.filter_by(status='unread').count(),
        recent_students=Student.query.order_by(Student.id.desc()).limit(5).all(),
        recent_certificates=Certificate.query.order_by(Certificate.id.desc()).limit(5).all(),
        recent_marksheets=Marksheet.query.order_by(Marksheet.id.desc()).limit(5).all(),
    )


# ============================================================
# STUDENTS MANAGEMENT
# ============================================================

@admin_bp.route('/students')
@protect
def students():
    q = clean(request.args.get('q'), 200)
    query = Student.query
    if q:
        query = query.filter(
            (Student.student_id.ilike(f'%{q}%')) |
            (Student.full_name.ilike(f'%{q}%')) |
            (Student.email.ilike(f'%{q}%'))
        )
    return render_template('admin/students.html', students=query.order_by(Student.id.desc()).all(), q=q)


@admin_bp.route('/students/add', methods=['GET', 'POST'])
@protect
def add_student():
    courses = Course.query.filter_by(status='active').order_by(Course.course_name).all()
    if request.method == 'POST':
        sid = clean(request.form.get('student_id'), 80)
        name = clean(request.form.get('full_name'), 200)
        email = clean(request.form.get('email'), 200)

        if not sid or not name or not email:
            flash('Student ID, name and email are required.', 'error')
            return redirect(url_for('admin.add_student'))
        if Student.query.filter_by(student_id=sid).first():
            flash('Student ID already exists.', 'error')
            return redirect(url_for('admin.add_student'))

        student = Student(
            student_id=sid,
            full_name=name,
            email=email,
            phone=clean(request.form.get('phone'), 50),
            gender=clean(request.form.get('gender'), 30),
            academic_session=clean(request.form.get('academic_session'), 50),
            admission_status=clean(request.form.get('admission_status'), 50) or 'active',
            completion_status=clean(request.form.get('completion_status'), 50) or 'ongoing',
        )

        course_id = request.form.get('course_id')
        if course_id:
            try:
                course = db.session.get(Course, int(course_id))
                if not course or course.status != 'active':
                    raise ValueError
                student.course_id = course.id
            except (ValueError, TypeError):
                flash('Invalid course selected.', 'error')
                return redirect(url_for('admin.add_student'))

        admission_date = request.form.get('admission_date')
        if admission_date:
            try:
                student.admission_date = datetime.strptime(admission_date, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid admission date.', 'error')
                return redirect(url_for('admin.add_student'))

        db.session.add(student)
        db.session.commit()
        audit('CREATE_STUDENT', sid)
        flash('Student added successfully.', 'success')
        return redirect(url_for('admin.students'))

    return render_template('admin/student_form.html', student=None, courses=courses)


@admin_bp.route('/students/<int:student_id>/edit', methods=['GET', 'POST'])
@protect
def edit_student(student_id):
    student = db.session.get(Student, student_id)
    if not student:
        flash('Student record not found.', 'error')
        return redirect(url_for('admin.students'))

    courses = Course.query.filter_by(status='active').order_by(Course.course_name).all()

    if request.method == 'POST':
        sid = clean(request.form.get('student_id'), 80)
        name = clean(request.form.get('full_name'), 200)
        email = clean(request.form.get('email'), 200)

        if not sid or not name or not email:
            flash('Student ID, name and email are required.', 'error')
            return redirect(url_for('admin.edit_student', student_id=student.id))

        existing = Student.query.filter_by(student_id=sid).first()
        if existing and existing.id != student.id:
            flash('Student ID already in use by another student.', 'error')
            return redirect(url_for('admin.edit_student', student_id=student.id))

        student.student_id = sid
        student.full_name = name
        student.email = email
        student.phone = clean(request.form.get('phone'), 50)
        student.gender = clean(request.form.get('gender'), 30)
        student.academic_session = clean(request.form.get('academic_session'), 50)
        student.admission_status = clean(request.form.get('admission_status'), 50) or 'active'
        student.completion_status = clean(request.form.get('completion_status'), 50) or 'ongoing'

        course_id = request.form.get('course_id')
        if course_id:
            try:
                course = db.session.get(Course, int(course_id))
                if not course:
                    raise ValueError
                student.course_id = course.id
            except (ValueError, TypeError):
                flash('Invalid course selected.', 'error')
                return redirect(url_for('admin.edit_student', student_id=student.id))
        else:
            student.course_id = None

        admission_date = request.form.get('admission_date')
        if admission_date:
            try:
                student.admission_date = datetime.strptime(admission_date, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid admission date.', 'error')
                return redirect(url_for('admin.edit_student', student_id=student.id))
        else:
            student.admission_date = None

        db.session.commit()
        audit('EDIT_STUDENT', f'{student.student_id} - {student.full_name}')
        flash('Student record updated successfully.', 'success')
        return redirect(url_for('admin.students'))

    return render_template('admin/student_form.html', student=student, courses=courses)


@admin_bp.route('/students/<int:student_id>/delete', methods=['POST'])
@protect
def delete_student(student_id):
    student = db.session.get(Student, student_id)
    if not student:
        flash('Student not found.', 'error')
        return redirect(url_for('admin.students'))

    sid = student.student_id
    name = student.full_name

    upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
    for cert in student.certificates:
        (upload_dir / cert.file_name).unlink(missing_ok=True)
    for mark in student.marksheets:
        (upload_dir / mark.file_name).unlink(missing_ok=True)

    db.session.delete(student)
    db.session.commit()
    audit('DELETE_STUDENT', f'{sid} - {name}')
    flash(f'Student {name} ({sid}) deleted successfully.', 'success')
    return redirect(url_for('admin.students'))


@admin_bp.route('/students/import', methods=['GET', 'POST'])
@protect
def import_students():
    if request.method == 'POST':
        uploaded = request.files.get('file')
        if not uploaded or not uploaded.filename or not uploaded.filename.lower().endswith('.xlsx'):
            flash('Please upload an .xlsx file.', 'error')
            return redirect(url_for('admin.import_students'))

        try:
            rows = list(load_workbook(uploaded, read_only=True, data_only=True).active.iter_rows(values_only=True))
        except Exception:
            flash('The uploaded spreadsheet could not be read.', 'error')
            return redirect(url_for('admin.import_students'))

        if not rows:
            flash('Spreadsheet is empty.', 'error')
            return redirect(url_for('admin.import_students'))

        headers = [str(x).strip().lower() if x is not None else '' for x in rows[0]]
        required = {'student_id', 'full_name', 'email'}
        if not required.issubset(headers):
            flash('Required columns: student_id, full_name, email', 'error')
            return redirect(url_for('admin.import_students'))

        count = skipped = 0
        for row in rows[1:]:
            data = dict(zip(headers, row))
            sid = clean(str(data.get('student_id') or ''), 80)
            name = clean(str(data.get('full_name') or ''), 200)
            email = clean(str(data.get('email') or ''), 200)
            if not sid or not name or not email or Student.query.filter_by(student_id=sid).first():
                skipped += 1
                continue

            db.session.add(Student(
                student_id=sid,
                full_name=name,
                email=email,
                phone=clean(str(data.get('phone') or ''), 50),
                gender=clean(str(data.get('gender') or ''), 30),
                academic_session=clean(str(data.get('academic_session') or ''), 50),
                admission_status=clean(str(data.get('admission_status') or 'active'), 50),
                completion_status=clean(str(data.get('completion_status') or 'ongoing'), 50),
            ))
            count += 1

        db.session.commit()
        audit('IMPORT_STUDENTS', f'{count} imported / {skipped} skipped')
        flash(f'Imported {count} students; skipped {skipped}.', 'success')
        return redirect(url_for('admin.students'))

    return render_template('admin/import_students.html')


# ============================================================
# COURSES MANAGEMENT
# ============================================================

@admin_bp.route('/courses')
@protect
def courses():
    return render_template('admin/courses.html', courses=Course.query.order_by(Course.id.desc()).all())


@admin_bp.route('/courses/add', methods=['GET', 'POST'])
@protect
def add_course():
    if request.method == 'POST':
        code = clean(request.form.get('course_code'), 50)
        name = clean(request.form.get('course_name'), 200)
        if not code or not name:
            flash('Course code and name are required.', 'error')
            return redirect(url_for('admin.add_course'))
        if Course.query.filter_by(course_code=code).first():
            flash('Course code already exists.', 'error')
            return redirect(url_for('admin.add_course'))

        course = Course(
            course_code=code,
            course_name=name,
            duration=clean(request.form.get('duration'), 100),
            eligibility=clean(request.form.get('eligibility'), 500),
            status=clean(request.form.get('status'), 30) or 'active',
            description=(request.form.get('description') or '').strip()[:10000],
        )
        db.session.add(course)
        db.session.commit()
        audit('CREATE_COURSE', code)
        flash('Course added.', 'success')
        return redirect(url_for('admin.courses'))

    return render_template('admin/course_form.html', course=None)


@admin_bp.route('/courses/<int:course_id>/edit', methods=['GET', 'POST'])
@protect
def edit_course(course_id):
    course = db.session.get(Course, course_id)
    if not course:
        flash('Course not found.', 'error')
        return redirect(url_for('admin.courses'))

    if request.method == 'POST':
        code = clean(request.form.get('course_code'), 50)
        name = clean(request.form.get('course_name'), 200)
        if not code or not name:
            flash('Course code and name are required.', 'error')
            return redirect(url_for('admin.edit_course', course_id=course.id))

        existing = Course.query.filter_by(course_code=code).first()
        if existing and existing.id != course.id:
            flash('Course code already in use by another course.', 'error')
            return redirect(url_for('admin.edit_course', course_id=course.id))

        course.course_code = code
        course.course_name = name
        course.duration = clean(request.form.get('duration'), 100)
        course.eligibility = clean(request.form.get('eligibility'), 500)
        course.status = clean(request.form.get('status'), 30) or 'active'
        course.description = (request.form.get('description') or '').strip()[:10000]

        db.session.commit()
        audit('EDIT_COURSE', code)
        flash('Course updated successfully.', 'success')
        return redirect(url_for('admin.courses'))

    return render_template('admin/course_form.html', course=course)


@admin_bp.route('/courses/<int:course_id>/delete', methods=['POST'])
@protect
def delete_course(course_id):
    course = db.session.get(Course, course_id)
    if not course:
        flash('Course not found.', 'error')
        return redirect(url_for('admin.courses'))

    code = course.course_code
    Student.query.filter_by(course_id=course.id).update({'course_id': None})
    db.session.delete(course)
    db.session.commit()
    audit('DELETE_COURSE', code)
    flash(f'Course {code} deleted successfully.', 'success')
    return redirect(url_for('admin.courses'))


# ============================================================
# CERTIFICATES MANAGEMENT
# ============================================================

@admin_bp.route('/certificates', methods=['GET', 'POST'])
@protect
def certificates():
    if request.method == 'POST':
        sid = clean(request.form.get('student_id'), 80)
        number = clean(request.form.get('certificate_number'), 100)
        pdf = request.files.get('certificate_file')
        student = Student.query.filter_by(student_id=sid).first()

        if not student or not number or not pdf or not pdf.filename.lower().endswith('.pdf') or not pdf_signature(pdf):
            flash('Valid Student ID, certificate number and a valid PDF are required.', 'error')
            return redirect(url_for('admin.certificates'))
        if Certificate.query.filter_by(certificate_number=number).first():
            flash('Certificate number already exists.', 'error')
            return redirect(url_for('admin.certificates'))

        filename = secure_filename(f'cert_{number}') + '.pdf'
        upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
        upload_dir.mkdir(parents=True, exist_ok=True)
        target = upload_dir / filename
        pdf.save(target)

        issue_date_val = request.form.get('issue_date')
        if issue_date_val:
            try:
                issue_date = datetime.strptime(issue_date_val, '%Y-%m-%d').date()
            except ValueError:
                issue_date = datetime.utcnow().date()
        else:
            issue_date = datetime.utcnow().date()

        try:
            certificate = Certificate(
                certificate_number=number,
                student_id=student.id,
                issue_date=issue_date,
                status='valid',
                file_name=filename,
            )
            db.session.add(certificate)
            student.certificate_number = number
            student.certificate_issue_date = issue_date
            db.session.commit()
        except Exception:
            db.session.rollback()
            target.unlink(missing_ok=True)
            flash('Certificate could not be saved.', 'error')
            return redirect(url_for('admin.certificates'))

        audit('UPLOAD_CERTIFICATE', number)
        flash('Certificate uploaded.', 'success')
        return redirect(url_for('admin.certificates'))

    return render_template('admin/certificates.html', certificates=Certificate.query.order_by(Certificate.id.desc()).all())


@admin_bp.route('/certificates/<int:certificate_id>/edit', methods=['GET', 'POST'])
@protect
def edit_certificate(certificate_id):
    certificate = db.session.get(Certificate, certificate_id)
    if not certificate:
        flash('Certificate not found.', 'error')
        return redirect(url_for('admin.certificates'))

    if request.method == 'POST':
        number = clean(request.form.get('certificate_number'), 100)
        status = clean(request.form.get('status'), 30) or 'valid'
        issue_date_str = request.form.get('issue_date')
        pdf = request.files.get('certificate_file')

        if not number:
            flash('Certificate number is required.', 'error')
            return redirect(url_for('admin.edit_certificate', certificate_id=certificate.id))

        existing = Certificate.query.filter_by(certificate_number=number).first()
        if existing and existing.id != certificate.id:
            flash('Certificate number already in use.', 'error')
            return redirect(url_for('admin.edit_certificate', certificate_id=certificate.id))

        if issue_date_str:
            try:
                certificate.issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid issue date format.', 'error')
                return redirect(url_for('admin.edit_certificate', certificate_id=certificate.id))

        certificate.certificate_number = number
        certificate.status = status

        if pdf and pdf.filename:
            if not pdf.filename.lower().endswith('.pdf') or not pdf_signature(pdf):
                flash('Uploaded file must be a valid PDF.', 'error')
                return redirect(url_for('admin.edit_certificate', certificate_id=certificate.id))

            upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
            upload_dir.mkdir(parents=True, exist_ok=True)

            old_path = upload_dir / certificate.file_name
            old_path.unlink(missing_ok=True)

            new_filename = secure_filename(f'cert_{number}_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}') + '.pdf'
            target = upload_dir / new_filename
            pdf.save(target)
            certificate.file_name = new_filename

        student = db.session.get(Student, certificate.student_id)
        if student:
            student.certificate_number = number
            student.certificate_issue_date = certificate.issue_date

        db.session.commit()
        audit('EDIT_CERTIFICATE', number)
        flash('Certificate updated successfully.', 'success')
        return redirect(url_for('admin.certificates'))

    return render_template('admin/certificate_form.html', certificate=certificate)


@admin_bp.route('/certificates/<int:certificate_id>/delete', methods=['POST'])
@protect
def delete_certificate(certificate_id):
    certificate = db.session.get(Certificate, certificate_id)
    if not certificate:
        flash('Certificate not found.', 'error')
        return redirect(url_for('admin.certificates'))

    number = certificate.certificate_number
    upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
    (upload_dir / certificate.file_name).unlink(missing_ok=True)

    student = db.session.get(Student, certificate.student_id)
    if student and student.certificate_number == certificate.certificate_number:
        student.certificate_number = None
        student.certificate_issue_date = None

    db.session.delete(certificate)
    db.session.commit()
    audit('DELETE_CERTIFICATE', number)
    flash(f'Certificate {number} deleted successfully.', 'success')
    return redirect(url_for('admin.certificates'))


# ============================================================
# MARKSHEETS MANAGEMENT
# ============================================================

@admin_bp.route('/marksheets', methods=['GET', 'POST'])
@protect
def marksheets():
    if request.method == 'POST':
        sid = clean(request.form.get('student_id'), 80)
        number = clean(request.form.get('marksheet_number'), 100)
        pdf = request.files.get('marksheet_file')
        student = Student.query.filter_by(student_id=sid).first()

        if not student or not number or not pdf or not pdf.filename.lower().endswith('.pdf') or not pdf_signature(pdf):
            flash('Valid Student ID, marksheet number and a valid PDF are required.', 'error')
            return redirect(url_for('admin.marksheets'))
        if Marksheet.query.filter_by(marksheet_number=number).first():
            flash('Marksheet number already exists.', 'error')
            return redirect(url_for('admin.marksheets'))

        filename = secure_filename(f'marksheet_{number}') + '.pdf'
        upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
        upload_dir.mkdir(parents=True, exist_ok=True)
        target = upload_dir / filename
        pdf.save(target)

        issue_date_val = request.form.get('issue_date')
        if issue_date_val:
            try:
                issue_date = datetime.strptime(issue_date_val, '%Y-%m-%d').date()
            except ValueError:
                issue_date = datetime.utcnow().date()
        else:
            issue_date = datetime.utcnow().date()

        try:
            marksheet = Marksheet(
                marksheet_number=number,
                student_id=student.id,
                issue_date=issue_date,
                status='valid',
                file_name=filename,
            )
            db.session.add(marksheet)
            student.marksheet_number = number
            student.marksheet_issue_date = issue_date
            db.session.commit()
        except Exception:
            db.session.rollback()
            target.unlink(missing_ok=True)
            flash('Marksheet could not be saved.', 'error')
            return redirect(url_for('admin.marksheets'))

        audit('UPLOAD_MARKSHEET', number)
        flash('Marksheet uploaded successfully.', 'success')
        return redirect(url_for('admin.marksheets'))

    return render_template('admin/marksheets.html', marksheets=Marksheet.query.order_by(Marksheet.id.desc()).all())


@admin_bp.route('/marksheets/<int:marksheet_id>/edit', methods=['GET', 'POST'])
@protect
def edit_marksheet(marksheet_id):
    marksheet = db.session.get(Marksheet, marksheet_id)
    if not marksheet:
        flash('Marksheet not found.', 'error')
        return redirect(url_for('admin.marksheets'))

    if request.method == 'POST':
        number = clean(request.form.get('marksheet_number'), 100)
        status = clean(request.form.get('status'), 30) or 'valid'
        issue_date_str = request.form.get('issue_date')
        pdf = request.files.get('marksheet_file')

        if not number:
            flash('Marksheet number is required.', 'error')
            return redirect(url_for('admin.edit_marksheet', marksheet_id=marksheet.id))

        existing = Marksheet.query.filter_by(marksheet_number=number).first()
        if existing and existing.id != marksheet.id:
            flash('Marksheet number already in use.', 'error')
            return redirect(url_for('admin.edit_marksheet', marksheet_id=marksheet.id))

        if issue_date_str:
            try:
                marksheet.issue_date = datetime.strptime(issue_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid issue date format.', 'error')
                return redirect(url_for('admin.edit_marksheet', marksheet_id=marksheet.id))

        marksheet.marksheet_number = number
        marksheet.status = status

        if pdf and pdf.filename:
            if not pdf.filename.lower().endswith('.pdf') or not pdf_signature(pdf):
                flash('Uploaded file must be a valid PDF.', 'error')
                return redirect(url_for('admin.edit_marksheet', marksheet_id=marksheet.id))

            upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
            upload_dir.mkdir(parents=True, exist_ok=True)

            old_path = upload_dir / marksheet.file_name
            old_path.unlink(missing_ok=True)

            new_filename = secure_filename(f'marksheet_{number}_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}') + '.pdf'
            target = upload_dir / new_filename
            pdf.save(target)
            marksheet.file_name = new_filename

        student = db.session.get(Student, marksheet.student_id)
        if student:
            student.marksheet_number = number
            student.marksheet_issue_date = marksheet.issue_date

        db.session.commit()
        audit('EDIT_MARKSHEET', number)
        flash('Marksheet updated successfully.', 'success')
        return redirect(url_for('admin.marksheets'))

    return render_template('admin/marksheet_form.html', marksheet=marksheet)


@admin_bp.route('/marksheets/<int:marksheet_id>/delete', methods=['POST'])
@protect
def delete_marksheet(marksheet_id):
    marksheet = db.session.get(Marksheet, marksheet_id)
    if not marksheet:
        flash('Marksheet not found.', 'error')
        return redirect(url_for('admin.marksheets'))

    number = marksheet.marksheet_number
    upload_dir = Path(current_app.config['UPLOAD_FOLDER']).resolve()
    (upload_dir / marksheet.file_name).unlink(missing_ok=True)

    student = db.session.get(Student, marksheet.student_id)
    if student and student.marksheet_number == marksheet.marksheet_number:
        student.marksheet_number = None
        student.marksheet_issue_date = None

    db.session.delete(marksheet)
    db.session.commit()
    audit('DELETE_MARKSHEET', number)
    flash(f'Marksheet {number} deleted successfully.', 'success')
    return redirect(url_for('admin.marksheets'))


# ============================================================
# MESSAGES, SECURITY LOGS, PASSWORDS, BACKUPS
# ============================================================

@admin_bp.route('/messages')
@protect
def messages():
    messages = ContactMessage.query.order_by(ContactMessage.id.desc()).all()
    unread = [m for m in messages if m.status == 'unread']
    for message in unread:
        message.status = 'read'
    if unread:
        db.session.commit()
    return render_template('admin/messages.html', messages=messages)


@admin_bp.route('/audit-logs')
@protect
def audit_logs():
    return render_template('admin/audit_logs.html', logs=AuditLog.query.order_by(AuditLog.id.desc()).limit(200).all())


@admin_bp.route('/change-password', methods=['GET', 'POST'])
@protect
def change_password():
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if not current_user.check_password(current):
            flash('Current password is incorrect.', 'error')
        elif len(new) < 12:
            flash('New password must be at least 12 characters.', 'error')
        elif new != confirm:
            flash('Passwords do not match.', 'error')
        else:
            current_user.set_password(new)
            db.session.commit()
            audit('CHANGE_PASSWORD', current_user.username)
            flash('Password changed. Please sign in again.', 'success')
            from flask_login import logout_user
            logout_user()
            return redirect(url_for('auth.login'))
    return render_template('admin/change_password.html')


@admin_bp.route('/backup', methods=['GET', 'POST'])
@protect
def backup():
    if request.method == 'POST':
        try:
            archive = create_backup(current_app)
            flash('Backup created successfully: ' + os.path.basename(archive), 'success')
            audit('CREATE_BACKUP', os.path.basename(archive))
        except Exception as e:
            flash(f'Backup could not be created: {str(e)}', 'error')

    folder = Path(current_app.config['BACKUP_FOLDER'])
    backups = []
    if folder.exists():
        for p in sorted(folder.glob('*.zip'), key=lambda x: x.stat().st_mtime, reverse=True):
            if p.is_file():
                size_kb = p.stat().st_size / 1024
                size_str = f"{size_kb / 1024:.2f} MB" if size_kb > 1024 else f"{size_kb:.1f} KB"
                mtime = datetime.fromtimestamp(p.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                backups.append({
                    'name': p.name,
                    'size': size_str,
                    'mtime': mtime
                })

    return render_template('admin/backup.html', backups=backups)


@admin_bp.route('/backup/download/<path:filename>')
@protect
def download_backup(filename):
    folder = Path(current_app.config['BACKUP_FOLDER']).resolve()
    path = (folder / filename).resolve()
    if folder not in path.parents or not path.is_file() or path.suffix.lower() != '.zip':
        return 'Backup not found', 404
    audit('DOWNLOAD_BACKUP', filename)
    return send_file(path, as_attachment=True, download_name=path.name)


@admin_bp.route('/backup/delete/<path:filename>', methods=['POST'])
@protect
def delete_backup(filename):
    folder = Path(current_app.config['BACKUP_FOLDER']).resolve()
    path = (folder / filename).resolve()
    if folder in path.parents and path.is_file() and path.suffix.lower() == '.zip':
        path.unlink(missing_ok=True)
        audit('DELETE_BACKUP', filename)
        flash(f'Backup archive {filename} deleted.', 'success')
    else:
        flash('Backup file not found.', 'error')
    return redirect(url_for('admin.backup'))

